import io
import uuid
from typing import Dict, List, Optional
import pandas as pd
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.ext.asyncio import AsyncSession
from exceptions.exception import ServiceException
from module_admin.dao.protocol_dao import ProtocolDao
from module_admin.entity.vo.common_vo import CrudResponseModel
from module_admin.entity.vo.protocol_vo import (
    ProtocolModel,
    ProtocolPageQueryModel,
    DeleteProtocolModel
)
from utils.common_util import CamelCaseUtil
from utils.page_util import PageResponseModel


PROTOCOL_TYPE_MAPPING = {
    'lan': 'LAN',
    'ethernet': 'LAN',
    'rs422': 'RS422',
    'rs485': 'RS485',
    'can': 'CAN',
    '1553b': '1553B',
}

PROTOCOL_DISPLAY_NAME = {
    'LAN': '以太网',
    'RS422': 'RS422',
    'RS485': 'RS485',
    'CAN': 'CAN',
    '1553B': '1553B',
}

# 不同协议类型对应的“报文配置”区域字段定义，中文名称需与前端协议模板中的标签保持一致
HEADER_FIELDS_BY_PROTOCOL = {
    'LAN': [
        ('sender', '发送方'),
        ('receiver', '接收方'),
        ('frequency', '传输频率'),
        ('speed', '传输速率/bps'),
        ('method', '传输方式'),
        ('port', '端口号'),
        ('frameLength', '帧长度/Byte'),
        ('errorHandling', '错误处理'),
    ],
    'RS422': [
        ('sender', '发送方'),
        ('receiver', '接收方'),
        ('frequency', '传输频率'),
        ('speed', '传输速率/bps'),
        ('method', '传输方式'),
        ('sendDuration', '发送时长/ms'),
        ('frameLength', '帧长度/Byte'),
        ('errorHandling', '错误处理'),
    ],
    'RS485': [
        ('sender', '发送方'),
        ('receiver', '接收方'),
        ('frequency', '传输频率'),
        ('speed', '传输速率/bps'),
        ('method', '传输方式'),
        ('sendDuration', '发送时长/ms'),
        ('frameLength', '帧长度/Byte'),
        ('errorHandling', '错误处理'),
    ],
    'CAN': [
        ('sender', '发送方'),
        ('receiver', '接收方'),
        ('frequency', '传输频率'),
        ('speed', '传输速率/bps'),
        ('method', '传输方式'),
        ('canId', 'ID'),
        ('frameLength', '帧长度/Byte'),
        ('errorHandling', '错误处理'),
    ],
    '1553B': [
        ('sender', '发送方'),
        ('receiver', '接收方'),
        ('frequency', '传输频率'),
        ('speed', '传输速率/bps'),
        ('method', '传输方式'),
        ('subAddress', '子地址'),
        ('frameLength', '帧长度/word'),
        ('errorHandling', '错误处理'),
    ],
}

# 通用头部字段定义，兼容旧模板
HEADER_FIELDS = [
    ('sender', '发送方'),
    ('receiver', '接收方'),
    ('frequency', '传输频率'),
    ('baudRate', '传输速率/bps'),
    ('method', '传输方式'),
    ('duration', '发送时长/ms'),
    ('frameLength', '帧长度/Byte'),
    ('errorHandling', '错误处理'),
]

HEADER_NUMERIC_KEYS = {'baudRate', 'duration', 'frameLength', 'port', 'sendDuration'}
HEADER_SHEET_NAMES = ['报文配置', '报文头', 'header', '基础配置', '协议配置表格']

FIELD_COLUMNS = [
    {'key': 'seq', 'label': '序号', 'width': 10},
    {'key': 'field_name', 'label': '信息名称', 'width': 24},
    {'key': 'byte_count', 'label': '字节数', 'width': 12},
    {'key': 'byte_sequence', 'label': '字节序号', 'width': 14},
    {'key': 'value_range', 'label': '值域及含义', 'width': 32},
    {'key': 'unit', 'label': '量纲', 'width': 12},
    {'key': 'data_type', 'label': '数据类型', 'width': 14},
    {'key': 'scale', 'label': '比例尺', 'width': 12},
    {'key': 'parent_code', 'label': '父级标识', 'width': 16},
    {'key': 'remark', 'label': '备注', 'width': 24},
]

FIELD_SHEET_NAMES = ['字段列表', '字段定义', 'fields', '报文字段', '字段列表表格']

FIELD_COLUMN_ALIASES: Dict[str, List[str]] = {
    'field_code': ['字段标识', '字段ID', '字段编号', 'fieldCode', 'field_id'],
    'field_name': ['信息名称', '字段名称', 'name', 'fieldName'],
    'byte_count': ['字节数', 'byteCount'],
    'byte_sequence': ['字节序号', 'byteSequence'],
    'value_range': ['值域及含义', '值域', 'valueRange', 'meaning'],
    'unit': ['量纲', '单位', 'unit'],
    'data_type': ['数据类型', '类型', 'dataType'],
    'scale': ['比例尺', 'scale'],
    'parent_code': ['父级标识', '父级ID', 'parentId', 'parentCode'],
    'remark': ['备注', 'remark', 'note'],
}

DATA_TYPE_OPTIONS = ['UINT8', 'UINT16', 'UINT32', 'INT8', 'INT16', 'INT32', 'FLOAT', 'DOUBLE', 'STRING']
DATA_TYPE_ALIASES = {
    'uint8': 'UINT8', 'u8': 'UINT8', 'byte': 'UINT8', '无符号8位': 'UINT8',
    'uint16': 'UINT16', 'u16': 'UINT16',
    'uint32': 'UINT32', 'u32': 'UINT32',
    'int8': 'INT8', 'i8': 'INT8',
    'int16': 'INT16', 'i16': 'INT16',
    'int32': 'INT32', 'i32': 'INT32',
    'float': 'FLOAT', 'f32': 'FLOAT', '单精度': 'FLOAT',
    'double': 'DOUBLE', 'dbl': 'DOUBLE', 'f64': 'DOUBLE', '双精度': 'DOUBLE',
    'string': 'STRING', 'str': 'STRING', '文本': 'STRING', '字符串': 'STRING'
}

DEFAULT_HEADER_SAMPLE = {
    'sender': '设备A',
    'receiver': '地面站',
    'frequency': '单次',
    'baudRate': 115200,
    'method': '422',
    'duration': 1000,
    'frameLength': 128,
    'errorHandling': '重传'
}

DEFAULT_HEADER_SAMPLE_EXTRA = {
    'port': '8080',
    'sendDuration': 1000,
    'subAddress': '01',
    'canId': '0x01',
}

DEFAULT_FIELD_SAMPLE = [
    {
        'seq': 1,
        'field_name': '同步码1',
        'byte_count': 1,
        'byte_sequence': '0',
        'value_range': '0xEB',
        'unit': '',
        'data_type': 'UINT8',
        'scale': '1',
        'parent_code': '',
        'remark': ''
    },
    {
        'seq': 2,
        'field_name': '同步码2',
        'byte_count': 1,
        'byte_sequence': '1',
        'value_range': '0x90',
        'unit': '',
        'data_type': 'UINT8',
        'scale': '1',
        'parent_code': '',
        'remark': ''
    },
    {
        'seq': 3,
        'field_name': '报文字节数',
        'byte_count': 2,
        'byte_sequence': '2~3',
        'value_range': '0x0100',
        'unit': '',
        'data_type': 'UINT16',
        'scale': '1',
        'parent_code': '',
        'remark': ''
    }
]

CONFIG_TABLE_LAYOUT = {
    'LAN': [
        [('sender', '发送方')],
        [('receiver', '接收方')],
        [('frequency', '传输频率'), ('speed', '传输速率/bps')],
        [('method', '传输方式'), ('port', '端口号')],
        [('frameLength', '帧长度/Byte'), ('errorHandling', '错误处理')],
    ],
    'RS422': [
        [('sender', '发送方')],
        [('receiver', '接收方')],
        [('frequency', '传输频率'), ('speed', '传输速率/bps')],
        [('method', '传输方式'), ('sendDuration', '发送时长/ms')],
        [('frameLength', '帧长度/Byte'), ('errorHandling', '错误处理')],
    ],
    'RS485': [
        [('sender', '发送方')],
        [('receiver', '接收方')],
        [('frequency', '传输频率'), ('speed', '传输速率/bps')],
        [('method', '传输方式'), ('sendDuration', '发送时长/ms')],
        [('frameLength', '帧长度/Byte'), ('errorHandling', '错误处理')],
    ],
    'CAN': [
        [('sender', '发送方')],
        [('receiver', '接收方')],
        [('frequency', '传输频率'), ('speed', '传输速率/bps')],
        [('method', '传输方式'), ('canId', 'ID')],
        [('frameLength', '帧长度/Byte'), ('errorHandling', '错误处理')],
    ],
    '1553B': [
        [('sender', '发送方')],
        [('receiver', '接收方')],
        [('frequency', '传输频率'), ('speed', '传输速率/bps')],
        [('method', '传输方式'), ('subAddress', '子地址')],
        [('frameLength', '帧长度/word'), ('errorHandling', '错误处理')],
    ],
}

HEADER_LABEL_TO_KEY = {
    '协议类型': 'protocolType',
    'ProtocolType': 'protocolType',
    'Protocol Type': 'protocolType',
    '发送方': 'sender',
    '接收方': 'receiver',
    '传输频率': 'frequency',
    '发送频率': 'frequency',
    '传输速率/bps': 'speed',
    '波特率': 'baudRate',
    '传输方式': 'method',
    '发送方式': 'method',
    '端口号': 'port',
    '发送时长/ms': 'sendDuration',
    '发送时长': 'sendDuration',
    'Duration': 'duration',
    '帧长度/Byte': 'frameLength',
    '帧长度/byte': 'frameLength',
    '帧长度/word': 'frameLength',
    '帧长度/Word': 'frameLength',
    '帧长': 'frameLength',
    '子地址': 'subAddress',
    'ID': 'canId',
    'Id': 'canId',
    '错误处理': 'errorHandling',
}

ALL_HEADER_KEYS = {
    key for key, _ in HEADER_FIELDS
} | {key for fields in HEADER_FIELDS_BY_PROTOCOL.values() for key, _ in fields} | {'protocolType'}


class ProtocolService:
    """
    协议管理模块服务层
    """

    @classmethod
    async def get_protocol_list_services(
        cls, query_db: AsyncSession, query_object: ProtocolPageQueryModel, is_page: bool = False
    ):
        """
        获取协议列表信息service

        :param query_db: orm对象
        :param query_object: 查询参数对象
        :param is_page: 是否开启分页
        :return: 协议列表信息对象
        """
        protocol_list_result = await ProtocolDao.get_protocol_list(query_db, query_object, is_page)
        if is_page:
            return PageResponseModel(
                **{
                    **protocol_list_result.model_dump(),
                    'rows': CamelCaseUtil.transform_result(protocol_list_result.rows)
                }
            )
        else:
            return CamelCaseUtil.transform_result(protocol_list_result)

    @classmethod
    async def add_protocol_services(cls, query_db: AsyncSession, page_object: ProtocolModel):
        """
        新增协议信息service

        :param query_db: orm对象
        :param page_object: 新增协议对象
        :return: 新增协议校验结果
        """
        try:
            await ProtocolDao.add_protocol_dao(query_db, page_object)
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='新增成功')
        except Exception as e:
            await query_db.rollback()
            raise e

    @classmethod
    async def edit_protocol_services(cls, query_db: AsyncSession, page_object: ProtocolModel):
        """
        编辑协议信息service

        :param query_db: orm对象
        :param page_object: 编辑协议对象
        :return: 编辑协议校验结果
        """
        try:
            await ProtocolDao.edit_protocol_dao(query_db, page_object)
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='更新成功')
        except Exception as e:
            await query_db.rollback()
            raise e

    @classmethod
    async def delete_protocol_services(cls, query_db: AsyncSession, page_object: DeleteProtocolModel):
        """
        删除协议信息service

        :param query_db: orm对象
        :param page_object: 删除协议对象
        :return: 删除协议校验结果
        """
        try:
            await ProtocolDao.delete_protocol_dao(query_db, page_object)
            await query_db.commit()
            return CrudResponseModel(is_success=True, message='删除成功')
        except Exception as e:
            await query_db.rollback()
            raise e

    @classmethod
    async def protocol_detail_services(cls, query_db: AsyncSession, protocol_id: int):
        """
        获取协议详细信息service

        :param query_db: orm对象
        :param protocol_id: 协议ID
        :return: 协议ID对应的信息
        """
        protocol = await ProtocolDao.get_protocol_detail_by_id(query_db, protocol_id)
        if protocol:
            result = CamelCaseUtil.transform_result(protocol)
        else:
            result = CamelCaseUtil.transform_result({})

        return result

    @classmethod
    async def get_protocol_import_template_services(
        cls,
        protocol_type: Optional[str] = None,
        protocolType: Optional[str] = None,
        **kwargs
    ):
        """
        获取协议导入模板excel内容
        :param protocol_type: 协议类型 (lan/rs422/rs485/can/1553b)
        """
        # 兼容不同的参数命名
        protocol_type = protocol_type or protocolType

        normalized_protocol_type = PROTOCOL_TYPE_MAPPING.get(protocol_type.lower() if protocol_type else '', None)
        header_fields = HEADER_FIELDS_BY_PROTOCOL.get(normalized_protocol_type, HEADER_FIELDS)
        display_name = PROTOCOL_DISPLAY_NAME.get(normalized_protocol_type, normalized_protocol_type or '通用')
        layout_rows = CONFIG_TABLE_LAYOUT.get(
            normalized_protocol_type,
            cls.__build_default_layout(header_fields)
        )

        workbook = Workbook()
        header_sheet = workbook.active
        header_sheet.title = '协议配置表格'

        # 空表头行，保持与前端所见格子一致（无标题）
        header_sheet.append(['', '', '', '', ''])
        header_sheet.freeze_panes = 'A2'

        first_data_row = 2
        last_row = first_data_row + len(layout_rows) - 1
        header_sheet.merge_cells(start_row=first_data_row, start_column=1, end_row=last_row, end_column=1)
        protocol_cell = header_sheet.cell(row=first_data_row, column=1, value=display_name)
        protocol_cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_offset, pair in enumerate(layout_rows):
            current_row = first_data_row + row_offset
            if pair:
                key1, label1 = pair[0]
                header_sheet.cell(row=current_row, column=2, value=label1).alignment = Alignment(horizontal='center', vertical='center')
                header_sheet.cell(row=current_row, column=3, value=cls.__header_sample_value(key1))
            if len(pair) > 1:
                key2, label2 = pair[1]
                header_sheet.cell(row=current_row, column=4, value=label2).alignment = Alignment(horizontal='center', vertical='center')
                header_sheet.cell(row=current_row, column=5, value=cls.__header_sample_value(key2))

            # 如果当前行只有 label1 / value1，也做横向合并
            if len(pair) == 1 or (len(pair) > 1 and not pair[1][0]):
                header_sheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)

        header_sheet.column_dimensions['A'].width = 12
        header_sheet.column_dimensions['B'].width = 18
        header_sheet.column_dimensions['C'].width = 24
        header_sheet.column_dimensions['D'].width = 18
        header_sheet.column_dimensions['E'].width = 24

        # 在同一张表下继续绘制字段列表表格，保持与示例一致
        field_table_start = last_row + 2
        for column_index, column_meta in enumerate(FIELD_COLUMNS, start=1):
            cell = header_sheet.cell(row=field_table_start, column=column_index, value=column_meta['label'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            header_sheet.column_dimensions[get_column_letter(column_index)].width = column_meta['width']

        for row_index, row_data in enumerate(DEFAULT_FIELD_SAMPLE, start=field_table_start + 1):
            for column_index, column_meta in enumerate(FIELD_COLUMNS, start=1):
                value = cls.__field_value(column_meta['key'], row_data, row_index - field_table_start)
                header_sheet.cell(row=row_index, column=column_index, value=value)

        # 同时保留单独的字段列表表，用于导入解析兼容
        field_sheet = workbook.create_sheet(title='字段列表表格')
        field_sheet.freeze_panes = 'A2'
        header_fill = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
        for column_index, column_meta in enumerate(FIELD_COLUMNS, start=1):
            cell = field_sheet.cell(row=1, column=column_index, value=column_meta['label'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            field_sheet.column_dimensions[get_column_letter(column_index)].width = column_meta['width']

        for row_index, row_data in enumerate(DEFAULT_FIELD_SAMPLE, start=2):
            for column_index, column_meta in enumerate(FIELD_COLUMNS, start=1):
                value = cls.__field_value(column_meta['key'], row_data, row_index - 1)
                field_sheet.cell(row=row_index, column=column_index, value=value)

        data_type_column_index = next(
            (index for index, column_meta in enumerate(FIELD_COLUMNS, start=1) if column_meta['key'] == 'data_type'),
            None
        )
        if data_type_column_index:
            data_type_column_letter = get_column_letter(data_type_column_index)

            dv_header = DataValidation(
                type='list',
                formula1=f"\"{','.join(DATA_TYPE_OPTIONS)}\"",
                allow_blank=True,
                showErrorMessage=True,
            )
            header_start_row = field_table_start + 1
            dv_header.add(f'{data_type_column_letter}{header_start_row}:{data_type_column_letter}1048576')
            header_sheet.add_data_validation(dv_header)

            dv_fields = DataValidation(
                type='list',
                formula1=f"\"{','.join(DATA_TYPE_OPTIONS)}\"",
                allow_blank=True,
                showErrorMessage=True,
            )
            dv_fields.add(f'{data_type_column_letter}2:{data_type_column_letter}1048576')
            field_sheet.add_data_validation(dv_fields)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    async def protocol_import_preview_services(cls, file: UploadFile):
        """
        解析导入的协议配置文件
        """
        if not file or not file.filename:
            raise ServiceException(message='请选择要导入的文件')
        suffix = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if suffix not in ('xlsx', 'xls', 'csv'):
            raise ServiceException(message='仅支持 .xlsx/.xls/.csv 格式的文件')

        content = await file.read()
        if not content:
            raise ServiceException(message='文件内容为空')

        sheet_map = cls.__read_workbook(content, suffix)
        protocol_type = cls.__extract_protocol_type(sheet_map)
        header = cls.__extract_header(sheet_map)
        fields = cls.__extract_fields(sheet_map)
        if not fields:
            raise ServiceException(message='文件中未解析到任何字段，请检查模板内容')
        return {'header': header, 'fields': fields, 'protocolType': protocol_type}

    @staticmethod
    def __read_workbook(content: bytes, suffix: str):
        try:
            if suffix in ('xlsx', 'xls'):
                return pd.read_excel(io.BytesIO(content), sheet_name=None, dtype=str)
            text_io = io.StringIO(content.decode('utf-8-sig'))
            df = pd.read_csv(text_io, dtype=str)
            return {'字段列表': df}
        except Exception as exc:
            raise ServiceException(message=f'解析导入文件失败: {exc}') from exc

    @staticmethod
    def __build_default_layout(header_fields: List[tuple]):
        layout: List[List[tuple]] = []
        buffer: List[tuple] = []
        for item in header_fields:
            buffer.append(item)
            if len(buffer) == 2:
                layout.append(buffer)
                buffer = []
        if buffer:
            layout.append(buffer)
        return layout

    @staticmethod
    def __header_sample_value(key: str):
        if key in DEFAULT_HEADER_SAMPLE_EXTRA:
            return DEFAULT_HEADER_SAMPLE_EXTRA[key]
        return DEFAULT_HEADER_SAMPLE.get(key, '')

    @staticmethod
    def __field_value(column_key: str, row_data: Dict, default_seq: int):
        if column_key == 'seq':
            return row_data.get('seq', default_seq)
        if column_key == 'parent_code':
            return row_data.get('parent_code') or row_data.get('parentId') or ''
        return row_data.get(column_key, '')

    @staticmethod
    def __normalize_header_value(key: str, value: Optional[str]):
        if key in HEADER_NUMERIC_KEYS:
            return ProtocolService.__safe_number(value)
        if value is None:
            return ''
        return str(value).strip()

    @staticmethod
    def __iter_label_value_pairs(row: pd.Series, columns):
        """
        遍历行数据，尝试提取「标签-值」对，兼容字段1/值1、label/value 等双列布局
        """
        column_names = [str(col).strip() for col in columns]
        row_dict = {str(col).strip(): row.get(col) for col in columns}
        candidate_pairs = [
            ('字段1', '值1'),
            ('字段2', '值2'),
            ('label1', 'value1'),
            ('label2', 'value2'),
            ('字段', '值'),
            ('名称', '取值'),
        ]
        for label_col, value_col in candidate_pairs:
            if label_col in row_dict and value_col in row_dict:
                label = str(row_dict[label_col]).strip()
                if label != '':
                    yield label, row_dict[value_col]

        # 兼容第一列为字段名称、第二列为取值的简单两列表格（但跳过协议名称列）
        if len(column_names) >= 2:
            first_col = column_names[0]
            if first_col not in {'协议', 'Protocol', '协议名', '协议名称', 'protocolName', 'protocol'}:
                label = str(row_dict.get(first_col, '')).strip()
                if label:
                    yield label, row_dict.get(column_names[1], '')

    @staticmethod
    def __extract_protocol_type(sheet_map: Dict[str, pd.DataFrame]):
        """
        从报文配置sheet中提取协议类型

        同时兼容两种布局：
        1. 横向：第一行是表头，存在「协议类型」列，其首行数据为协议类型
        2. 纵向：第一列是字段名称，某一行的名称为「协议类型」，第二列为对应取值
        """
        config_df: Optional[pd.DataFrame] = None
        for sheet_name in HEADER_SHEET_NAMES:
            if sheet_name in sheet_map:
                config_df = sheet_map[sheet_name]
                break

        if config_df is None or config_df.empty:
            return None

        df = config_df.fillna('')
        col_aliases = {'协议类型', 'ProtocolType', 'Protocol Type'}
        display_reverse = {v.upper(): k.lower() for k, v in PROTOCOL_DISPLAY_NAME.items()}
        type_reverse_map = {v.upper(): k for k, v in PROTOCOL_TYPE_MAPPING.items()}

        # 方案一：横向表头，查找名为“协议类型”的列
        for column in df.columns:
            if str(column).strip() in col_aliases:
                series = df[column].astype(str).str.strip()
                for value in series:
                    if value:
                        protocol_type_raw = str(value).strip().upper()
                        return type_reverse_map.get(protocol_type_raw, protocol_type_raw.lower())

        # 方案二：纵向字段行，第一列作为字段名称，第二列存放值
        if len(df.columns) >= 2:
            name_col = df.columns[0]
            value_col = df.columns[1]
            for _, row in df.iterrows():
                field_name = str(row.get(name_col, '')).strip()
                if field_name in col_aliases:
                    value = str(row.get(value_col, '')).strip()
                    if value:
                        protocol_type_raw = value.upper()
                        return type_reverse_map.get(protocol_type_raw, protocol_type_raw.lower())

        # 方案三：字段/值双列布局（字段1/值1 等）
        for _, row in df.iterrows():
            for label, value in ProtocolService.__iter_label_value_pairs(row, df.columns):
                if str(label).strip() in col_aliases and value:
                    protocol_type_raw = str(value).strip().upper()
                    return type_reverse_map.get(protocol_type_raw, protocol_type_raw.lower())

        # 方案四：使用第一列的协议名称推断
        first_col_name = str(df.columns[0]).strip()
        if first_col_name in {'协议', 'Protocol', '协议名', '协议名称', 'protocolName', 'protocol'}:
            for _, row in df.iterrows():
                name_value = str(row.get(first_col_name, '')).strip()
                if not name_value:
                    continue
                mapped = display_reverse.get(name_value.upper())
                if mapped:
                    return mapped

        return None

    @classmethod
    def __extract_header(cls, sheet_map: Dict[str, pd.DataFrame]):
        header_df: Optional[pd.DataFrame] = None
        for sheet_name in HEADER_SHEET_NAMES:
            if sheet_name in sheet_map:
                header_df = sheet_map[sheet_name]
                break
        if header_df is None:
            first_sheet = next(iter(sheet_map.values()))
            if set(first_sheet.columns).issuperset({label for _, label in HEADER_FIELDS}):
                header_df = first_sheet
        header_result = {
            key: (None if key in HEADER_NUMERIC_KEYS else '')
            for key in ALL_HEADER_KEYS
        }
        if header_df is None or header_df.empty:
            return header_result

        df = header_df.fillna('')

        # 方案一：横向表头（列名为中文标签）
        row_data = df.iloc[0].to_dict()
        if any(label in row_data for label in HEADER_LABEL_TO_KEY.keys()):
            for label, key in HEADER_LABEL_TO_KEY.items():
                if label in row_data:
                    value = row_data.get(label, '')
                    header_result[key] = cls.__normalize_header_value(key, value)
            return header_result

        # 方案二：字段/值成对布局
        for _, row in df.iterrows():
            for label, value in cls.__iter_label_value_pairs(row, df.columns):
                key = HEADER_LABEL_TO_KEY.get(str(label).strip())
                if key is None:
                    continue
                header_result[key] = cls.__normalize_header_value(key, value)
        return header_result

    @classmethod
    def __extract_fields(cls, sheet_map: Dict[str, pd.DataFrame]):
        fields_sheet: Optional[pd.DataFrame] = None
        for sheet_name in FIELD_SHEET_NAMES:
            if sheet_name in sheet_map:
                fields_sheet = sheet_map[sheet_name]
                break
        if fields_sheet is None:
            candidate = next(iter(sheet_map.values()))
            if any(header in candidate.columns for header in FIELD_COLUMN_ALIASES.get('field_name', [])):
                fields_sheet = candidate
        if fields_sheet is None:
            raise ServiceException(message='未找到包含字段定义的工作表，请确认模板未被修改')

        records = fields_sheet.fillna('').to_dict(orient='records')
        resolved_fields = []
        code_lookup: Dict[str, str] = {}
        name_lookup: Dict[str, str] = {}
        for index, row in enumerate(records):
            normalized = cls.__normalize_field_row(row, index, resolved_fields, code_lookup, name_lookup)
            if normalized:
                resolved_fields.append(normalized)

        return resolved_fields

    @classmethod
    def __normalize_field_row(cls, row: Dict, index: int, resolved_fields: List[Dict], code_lookup: Dict[str, str], name_lookup: Dict[str, str]):
        normalized_row = {str(k).strip(): row.get(k) for k in row.keys()}
        field_name = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('field_name', []))
        field_code = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('field_code', []))
        if not field_name and not field_code:
            return None

        unique_id = str(field_code).strip() if field_code else cls.__generate_field_id(index)
        unique_id = str(unique_id).strip() or cls.__generate_field_id(index)
        byte_count = cls.__to_int(cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('byte_count', [])), default=1)
        byte_sequence = cls.__format_sequence(cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('byte_sequence', [])))
        value_range = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('value_range', [])) or ''
        unit = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('unit', [])) or ''
        data_type_raw = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('data_type', []))
        scale = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('scale', [])) or '1'
        remark = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('remark', [])) or ''
        parent_ref = cls.__pick_value(normalized_row, FIELD_COLUMN_ALIASES.get('parent_code', []))
        parent_id = cls.__resolve_parent_id(parent_ref, resolved_fields, code_lookup, name_lookup)

        if field_name:
            name_lookup[str(field_name).strip()] = unique_id
        if field_code:
            code_lookup[str(field_code).strip()] = unique_id

        return {
            'id': unique_id,
            'parentId': parent_id,
            'fieldName': str(field_name).strip() if field_name else '',
            'byteCount': byte_count,
            'byteSequence': str(byte_sequence).strip(),
            'valueRange': str(value_range).strip(),
            'unit': str(unit).strip(),
            'dataType': cls.__normalize_data_type(data_type_raw),
            'scale': str(scale).strip() or '1',
            'remark': str(remark).strip()
        }

    @staticmethod
    def __pick_value(row: Dict, keys: List[str]):
        for key in keys:
            value = row.get(key)
            if value is None:
                continue
            text = str(value).strip()
            if text != '':
                return value
        return None

    @staticmethod
    def __generate_field_id(index: int):
        return f'field_{index}_{uuid.uuid4().hex}'

    @staticmethod
    def __normalize_data_type(value: Optional[str]):
        if not value:
            return 'STRING'
        value_str = str(value).strip()
        upper = value_str.upper()
        if upper in DATA_TYPE_OPTIONS:
            return upper
        return DATA_TYPE_ALIASES.get(value_str.lower(), 'STRING')

    @staticmethod
    def __safe_number(value: Optional[str]):
        if value is None:
            return None
        text = str(value).strip()
        if text == '':
            return None
        try:
            number = float(text)
            return int(number) if number.is_integer() else number
        except ValueError:
            return None

    @classmethod
    def __to_int(cls, value: Optional[str], default: int = 1):
        number = cls.__safe_number(value)
        return default if number is None else int(number)

    @classmethod
    def __resolve_parent_id(cls, parent_ref: Optional[str], resolved_fields: List[Dict], code_lookup: Dict[str, str], name_lookup: Dict[str, str]):
        if parent_ref is None:
            return None
        ref_text = str(parent_ref).strip()
        if not ref_text:
            return None
        if ref_text in code_lookup:
            return code_lookup[ref_text]
        if ref_text in name_lookup:
            return name_lookup[ref_text]
        if ref_text.isdigit():
            ref_index = int(ref_text)
            if 0 <= ref_index < len(resolved_fields):
                return resolved_fields[ref_index]['id']
            if 1 <= ref_index <= len(resolved_fields):
                return resolved_fields[ref_index - 1]['id']
        return None

    @classmethod
    def __format_sequence(cls, value: Optional[str]):
        if value is None:
            return ''
        number = cls.__safe_number(value)
        if number is None:
            return str(value).strip()
        if isinstance(number, float) and number.is_integer():
            return str(int(number))
        return str(number)
